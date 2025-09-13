import argparse
import logging
import os
import sys
import pandas as pd

from openpyxl import load_workbook

# Add parent directory to path for imports
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from Checklist.stock import fetch_and_score_stock
from Checklist.management import fetch_and_score_management, clear_mission_statements_file
from Checklist.financial import fetch_and_score_financials
from Checklist.potential import fetch_and_score_potential
from Checklist.customers import fetch_and_score_customers
from Checklist.penalties import fetch_and_score_penalties
from Checklist.specific_factors import fetch_and_score_specific_factors
from Checklist.settings import EXCEL_NAME, PORTFOLIO_SHEET_NAME, FEROLDI_SHEET_NAME, FINANCIAL_ROWS, POTENTIAL_ROWS, \
    CUSTOMERS_ROWS, MANAGEMENT_ROWS, STOCK_ROWS, PENALTIES_ROWS, SPECIFIC_FACTORS_ROWS
from Checklist.utils import setup_logging
from Checklist.utilities.logging_config import get_logger, log_data_issue, log_debug

# Setup logging
setup_logging()

SECTION_ROW_MAPS = {
    "financial": FINANCIAL_ROWS,
    "potential": POTENTIAL_ROWS,
    "customers": CUSTOMERS_ROWS,
    "management": MANAGEMENT_ROWS,
    "stock": STOCK_ROWS,
    "penalties": PENALTIES_ROWS,
    "specific_factors": SPECIFIC_FACTORS_ROWS
}

# Section configuration mapping
SECTION_CONFIG = {
    "financial": {
        "function": fetch_and_score_financials,
        "emoji": "📊",
        "name": "Financial data"
    },
    "potential": {
        "function": fetch_and_score_potential,
        "emoji": "🚀",
        "name": "Potential data"
    },
    "customers": {
        "function": fetch_and_score_customers,
        "emoji": "👥",
        "name": "Customers data"
    },
    "specific_factors": {
        "function": fetch_and_score_specific_factors,
        "emoji": "🎯",
        "name": "Specific factors"
    },
    "management": {
        "function": fetch_and_score_management,
        "emoji": "👔",
        "name": "Management data"
    },
    "stock": {
        "function": fetch_and_score_stock,
        "emoji": "📈",
        "name": "Stock data"
    },
    "penalties": {
        "function": fetch_and_score_penalties,
        "emoji": "⚠️",
        "name": "Penalties"
    }
}


def clear_management_file_once():
    """Helper to clear management file with error handling."""
    try:
        clear_mission_statements_file()
    except Exception as e:
        logging.error(f"Error clearing mission statements file: {e}")


def process_sections_for_ticker(ticker, sections_to_run, show_progress=False, indent=""):
    """
    Process all specified sections for a single ticker.
    
    Args:
        ticker (str): The stock ticker to process
        sections_to_run (list): List of sections to run
        show_progress (bool): Whether to show progress messages
        indent (str): Indentation for progress messages
    
    Returns:
        dict: Results dictionary with section data
    """
    results_dict = {}
    
    for section in sections_to_run:
        if section not in SECTION_CONFIG:
            log_data_issue(ticker, f"Unknown section: {section}")
            continue
            
        config = SECTION_CONFIG[section]
        
        if show_progress:
            print(f"{indent}{config['emoji']} {config['name']}...")
        
        # Special handling for management section
        if section == "management":
            # Management clearing is handled externally to avoid multiple clears
            pass
            
        results_dict[section] = run_section([ticker], config["function"], section)
    
    return results_dict


def process_all_sections_batch(tickers, sections_to_run):
    """
    Process all sections for multiple tickers using the original batch method.
    This is more efficient for large batches as it processes each section for all tickers together.
    
    Args:
        tickers (list): List of tickers to process
        sections_to_run (list): List of sections to run
    
    Returns:
        dict: Results dictionary organized by section
    """
    results_dict = {}
    
    # Handle management section clearing once at the beginning
    if "management" in sections_to_run:
        clear_management_file_once()
    
    for section in sections_to_run:
        if section not in SECTION_CONFIG:
            log_data_issue("SYSTEM", f"Unknown section: {section}")
            continue
            
        config = SECTION_CONFIG[section]
        results_dict[section] = run_section(tickers, config["function"], section)
    
    return results_dict


def load_tickers_from_excel(file_path, sheet_name):
    try:
        log_debug(f"Loading tickers from '{file_path}' (sheet: '{sheet_name}').")
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        tickers = []
        for ticker in df["Ticker"].dropna().tolist():  # Assuming 'Ticker' column exists
            ticker = ticker.strip()
            if not ticker or len(ticker) > 5:
                log_debug(f"Invalid ticker '{ticker}' detected. Stopping ticker extraction.")
                break  # Stop processing
            tickers.append(ticker)

        return tickers
    except Exception as e:
        logging.error(f"Error loading tickers from Excel: {e}")
        return []


def update_excel(results_dict, excel_path, sheet_name):
    """
    Update financial scores in the specified Excel sheet.
    Only updates manual fields if the new score is not None
    """
    try:
        log_debug(f"Updating scores in '{excel_path}' (sheet: '{sheet_name}').")
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]

        for section, results in results_dict.items():
            rows_mapping = SECTION_ROW_MAPS.get(section, {})

            for col_idx, (ticker, scores) in enumerate(results.items(), start=4):
                if scores is None:
                    log_data_issue(ticker, f"Skipping {ticker} due to missing scores.")
                    continue

                log_debug(f"Updating {section} scores for {ticker} in column {col_idx}.")
                try:
                    for metric, row_idx in rows_mapping.items():
                        new_score = scores.get(metric, 0)
                        
                        # Only update if the new score is not None (allow 0 scores)
                        if new_score is not None:
                            sheet.cell(row=row_idx, column=col_idx, value=new_score)
                            log_debug(f"Updated {metric} for {ticker}: {new_score}")
                except Exception as e:
                    logging.error(f"Error updating scores for {ticker}: {e}")

        wb.save(excel_path)
        log_debug(f"Scores successfully updated in '{excel_path}'.")
    except Exception as e:
        logging.error(f"Error updating Excel file: {e}")


def run_section(tickers, fetch_function, section):
    results = {}
    for ticker in tickers:
        log_debug(f"Processing {section} data for ticker: {ticker}")
        try:
            scores = fetch_function(ticker)
            if scores:
                results[ticker] = scores
        except Exception as e:
            logging.error(f"Error processing {section} data for {ticker}: {e}")
            results[ticker] = None
    return results


def calculate_total_score(results_dict, ticker):
    """
    Calculate the total score for a ticker across all sections.
    """
    total_score = 0
    section_scores = {}
    
    log_debug(f"Calculating total score for {ticker}")
    
    for section, results in results_dict.items():
        if ticker in results and results[ticker] is not None:
            # Log the individual scores before summing
            log_debug(f"  {section} scores for {ticker}: {results[ticker]}")
            
            # Special handling for penalties section to avoid double counting the 'Gauntlet Score'
            if section == "penalties":
                # Exclude 'Gauntlet Score' to prevent double counting
                individual_scores = {k: v for k, v in results[ticker].items() if k != 'Gauntlet Score'}
                section_total = sum(score for score in individual_scores.values() if score is not None)
            else:
                section_total = sum(score for score in results[ticker].values() if score is not None)
                
            section_scores[section] = section_total
            total_score += section_total
            log_debug(f"  {section} total: {section_total}")
        else:
            section_scores[section] = 0
            log_debug(f"  {section}: No scores or None result")
    
    log_debug(f"Final total score for {ticker}: {total_score}")
    log_debug(f"Section breakdown: {section_scores}")
    
    return total_score, section_scores


def print_score_summary(ticker, total_score, section_scores):
    """
    Print a visually appealing score summary for the ticker.
    """
    print("\n" + "="*60)
    print(f"           SCORE SUMMARY FOR {ticker.upper()}")
    print("="*60)
    print()
    
    for section, score in section_scores.items():
        section_name = section.replace('_', ' ').title()
        print(f"{section_name:<20}: {score:>6}")
    
    print("-" * 30)
    print(f"{'TOTAL SCORE':<20}: {total_score:>6}")
    print("="*60)
    print()


def print_multiple_stocks_summary(stocks_results):
    """
    Print a summary table for multiple stocks showing all their scores.
    
    Args:
        stocks_results (dict): Dictionary with ticker as key and (total_score, section_scores) as values
    """
    if not stocks_results:
        print("No results to display.")
        return
    
    # Get all section names from the first stock
    first_ticker = list(stocks_results.keys())[0]
    section_names = list(stocks_results[first_ticker][1].keys())
    
    print("\n" + "="*120)
    print("                                    PORTFOLIO ANALYSIS SUMMARY")
    print("="*120)
    print()
    
    # Create header with proper spacing
    header = f"{'TICKER':<8}"
    for section in section_names:
        # Truncate and format section names to fit in 10 characters
        section_display = section.replace('_', ' ').title()
        if len(section_display) > 10:
            section_display = section_display[:10]
        header += f"{section_display:>10}"
    header += f"{'TOTAL':>10}"
    
    print(header)
    print("-" * len(header))
    
    # Print each stock's scores
    for ticker, (total_score, section_scores) in stocks_results.items():
        row = f"{ticker:<8}"
        for section in section_names:
            score = section_scores.get(section, 0)
            row += f"{score:>10}"
        row += f"{total_score:>10}"
        print(row)
    
    print("-" * len(header))
    
    # Calculate and display averages
    if len(stocks_results) > 1:
        avg_total = sum(total for total, _ in stocks_results.values()) / len(stocks_results)
        avg_row = f"{'AVG':<8}"
        
        for section in section_names:
            avg_section = sum(section_scores.get(section, 0) for _, section_scores in stocks_results.values()) / len(stocks_results)
            avg_row += f"{avg_section:>10.1f}"
        
        avg_row += f"{avg_total:>10.1f}"
        print(avg_row)
    
    print("="*120)
    print()
    
    # Print top performers
    if len(stocks_results) > 1:
        sorted_stocks = sorted(stocks_results.items(), key=lambda x: x[1][0], reverse=True)
        print("🏆 TOP PERFORMERS:")
        for i, (ticker, (total_score, _)) in enumerate(sorted_stocks[:3], 1):
            print(f"   {i}. {ticker}: {total_score} points")
        print()
    elif len(stocks_results) == 1:
        ticker, (total_score, _) = list(stocks_results.items())[0]
        print(f"📊 ANALYSIS COMPLETE FOR {ticker}: {total_score} points")
        print()


def analyze_single_stock(ticker, sections_to_run=None):
    """
    Run the entire analysis flow for a single stock and display the total score.
    This is for research purposes only - does not update Excel.
    
    Args:
        ticker (str): The stock ticker to analyze
        sections_to_run (list): List of sections to run, defaults to all sections
    
    Returns:
        tuple: (total_score, section_scores, results_dict)
    """
    if sections_to_run is None:
        sections_to_run = ["financial", "potential", "customers", "specific_factors", "management", "stock", "penalties"]
    
    print(f"\n🔍 Starting analysis for {ticker.upper()}...")
    print("-" * 40)
    
    # Handle management section clearing once at the beginning
    if "management" in sections_to_run:
        clear_management_file_once()
    
    # Process all sections using the helper method
    results_dict = process_sections_for_ticker(ticker, sections_to_run, show_progress=True)
    
    # Calculate and display total score
    total_score, section_scores = calculate_total_score(results_dict, ticker)
    
    # Create summary for single stock and display in table format
    stocks_summary = {ticker: (total_score, section_scores)}
    print_multiple_stocks_summary(stocks_summary)
    
    return total_score, section_scores, results_dict


def analyze_multiple_stocks(tickers, sections_to_run=None):
    """
    Run the entire analysis flow for multiple stocks and display a summary table.
    This is for research purposes only - does not update Excel.
    
    Args:
        tickers (list): List of stock tickers to analyze
        sections_to_run (list): List of sections to run, defaults to all sections
    
    Returns:
        dict: Dictionary with ticker as key and (total_score, section_scores, results_dict) as values
    """
    if sections_to_run is None:
        sections_to_run = ["financial", "potential", "customers", "specific_factors", "management", "stock", "penalties"]
    
    all_results = {}
    stocks_summary = {}
    
    print(f"\n🔍 Starting batch analysis for {len(tickers)} stocks...")
    print("="*60)
    
    # Handle management section clearing once at the beginning
    if "management" in sections_to_run:
        clear_management_file_once()
    
    # Process each ticker
    for i, ticker in enumerate(tickers, 1):
        print(f"\n📊 Processing {ticker.upper()} ({i}/{len(tickers)})...")
        print("-" * 40)
        
        # Process all sections for this ticker using the helper method
        results_dict = process_sections_for_ticker(ticker, sections_to_run, show_progress=True, indent="  ")
        
        # Calculate scores for this ticker
        total_score, section_scores = calculate_total_score(results_dict, ticker)
        
        # Store results
        all_results[ticker] = (total_score, section_scores, results_dict)
        stocks_summary[ticker] = (total_score, section_scores)
        
        print(f"  ✅ {ticker} completed - Total Score: {total_score}")
    
    # Display summary table for all stocks
    print_multiple_stocks_summary(stocks_summary)
    
    return all_results


def str_to_bool(v):
    """Convert string to boolean for argparse."""
    if isinstance(v, bool):
        return v
    if v.lower() in ('yes', 'true', 't', 'y', '1'):
        return True
    elif v.lower() in ('no', 'false', 'f', 'n', '0'):
        return False
    else:
        raise argparse.ArgumentTypeError('Boolean value expected.')


def parse_arguments():
    parser = argparse.ArgumentParser(description="Run financial and potential analysis on portfolio tickers.")
    parser.add_argument(
        "--stock",
        type=str,
        default=None,
        help="Single stock ticker to process. If not provided, tickers are loaded from the Portfolio sheet."
    )
    parser.add_argument(
        "--analyze",
        type=str,
        default=None,
        help="Single stock ticker to analyze with summary table display (research only - no Excel update)."
    )
    parser.add_argument(
        "--batch",
        type=str,
        nargs='+',
        default=None,
        help="List of stock tickers to analyze with summary table display (research only - no Excel update). Example: --batch AAPL MSFT GOOGL"
    )
    parser.add_argument(
        "--update_excel",
        type=str_to_bool,
        default=True,
        help="Flag to update the Feroldi Quality Score sheet with results (default: True). Only applies to main portfolio analysis and --stock option."
    )
    parser.add_argument(
        "--sections",
        type=str,
        nargs='+',
        default=["financial", "potential", "customers", "specific_factors", "management", "stock", "penalties"],
        help="By default we run all sections."
    )
    return parser.parse_args()


def main():
    log_debug("Starting main portfolio analysis run.")
    excel_path = os.path.join(os.getcwd(), f"{EXCEL_NAME}.xlsx")
    args = parse_arguments()

    stock = args.stock
    analyze_stock = args.analyze
    batch_stocks = args.batch
    update_excel_flag = args.update_excel
    sections_to_run = args.sections

    # If batch flag is used, run the multiple stocks analysis method (research only)
    if batch_stocks:
        all_results = analyze_multiple_stocks(
            batch_stocks,
            sections_to_run
        )
        log_debug("Batch analysis complete.")
        return

    # If analyze flag is used, run the single stock analysis method (research only)
    if analyze_stock:
        total_score, section_scores, results_dict = analyze_single_stock(
            analyze_stock, 
            sections_to_run
        )
        log_debug("Single stock analysis complete.")
        return

    # Original logic for batch processing or single stock with Excel update capability
    if stock:
        tickers = [stock]
        log_debug(f"Processing single stock: {stock}")
    else:
        tickers = load_tickers_from_excel(excel_path, PORTFOLIO_SHEET_NAME)

    # Use the batch processing helper method
    results_dict = process_all_sections_batch(tickers, sections_to_run)

    # Create summary table for consistent display
    stocks_summary = {}
    for ticker in tickers:
        total_score, section_scores = calculate_total_score(results_dict, ticker)
        stocks_summary[ticker] = (total_score, section_scores)
    
    # Display results in the same table format
    print_multiple_stocks_summary(stocks_summary)

    # Update Excel if requested
    if update_excel_flag:
        log_debug(f"Updating Excel with results: {results_dict}")
        update_excel(results_dict, excel_path, FEROLDI_SHEET_NAME)
        print("✅ Excel file updated with results")

    log_debug("Processed Results:")
    for section, results in results_dict.items():
        for ticker, scores in results.items():
            log_debug(f"{ticker} ({section}): {scores}")
    log_debug("Main portfolio analysis run complete.")


if __name__ == "__main__":
    main()
